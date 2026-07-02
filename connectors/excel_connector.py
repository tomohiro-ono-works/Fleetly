import os
import posixpath
import time
import zipfile
from datetime import datetime, date
from typing import Any
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

from connectors.base_connector import BaseConnector


XLSX_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
XLSX_OFFICE_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
XLSX_PACKAGE_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
DEFAULT_CHUNK_SIZE = 50000


class ExcelConnector(BaseConnector):
    def _resolve_usecols_from_schema_origin(self, schema: Any):
        if schema is None or str(schema).strip() == "":
            return None
        items = self.parse_schema_definition(schema)
        if not isinstance(items, list):
            return None
        usecols = []
        for item in items:
            if not isinstance(item, dict):
                continue
            origin_name = str(item.get("origin_name") or "").strip()
            if origin_name and origin_name not in usecols:
                usecols.append(origin_name)
        return usecols or None

    @staticmethod
    def _resolve_chunk_size(value: Any) -> int:
        if value is None:
            return DEFAULT_CHUNK_SIZE
        text = str(value).strip()
        if not text:
            return DEFAULT_CHUNK_SIZE
        chunk_size = int(text)
        if chunk_size <= 0:
            raise ValueError("chunk_size は 1 以上で指定してください。")
        return chunk_size

    @staticmethod
    def _normalize_excel_value(value: Any) -> Any:
        if pd.isna(value):
            return None
        if isinstance(value, datetime):
            normalized = value.replace(tzinfo=None) if value.tzinfo is not None else value
            iso_str = normalized.isoformat()
            return iso_str[:10] if iso_str.endswith("T00:00:00") else iso_str
        if isinstance(value, date):
            return value.isoformat()
        return value

    def preview_excel(self, path: str, sheet_name: str | None = None, max_rows: int = 30) -> dict[str, Any]:
        normalized_path = self.normalize_file_path(path)
        if normalized_path is None or not os.path.exists(normalized_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {normalized_path}")
        if not str(normalized_path).lower().endswith(".xlsx"):
            raise ValueError("Excel preview は .xlsx のみ対応しています。")

        row_limit = max(1, int(max_rows or 30))
        with zipfile.ZipFile(normalized_path) as archive:
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            sheets = [
                {
                    "name": sheet.attrib.get("name", ""),
                    "rid": sheet.attrib.get(XLSX_OFFICE_REL_NS + "id", ""),
                }
                for sheet in workbook.findall(".//" + XLSX_MAIN_NS + "sheet")
            ]
            sheet_names = [sheet["name"] for sheet in sheets if sheet["name"]]
            if not sheet_names:
                raise ValueError("シートがありません。")

            rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            targets = {}
            for rel in rels.findall(XLSX_PACKAGE_REL_NS + "Relationship"):
                target = str(rel.attrib.get("Target", "")).replace("\\", "/")
                targets[rel.attrib.get("Id", "")] = posixpath.normpath(
                    target.lstrip("/") if target.startswith("/") else posixpath.join("xl", target)
                )

            selected_sheet = str(sheet_name or "").strip()
            target_sheet_name = selected_sheet if selected_sheet in sheet_names else sheet_names[0]
            target_sheet = next(sheet for sheet in sheets if sheet.get("name") == target_sheet_name)
            worksheet_path = targets.get(target_sheet.get("rid", ""))
            if not worksheet_path:
                raise ValueError(f"シート参照が見つかりませんでした: {target_sheet_name}")
            if worksheet_path not in archive.namelist():
                raise FileNotFoundError(f"シート実体が見つかりません: {worksheet_path}")

            rows: list[list[Any]] = []
            shared_refs: set[int] = set()
            with archive.open(worksheet_path) as handle:
                for _event, row_elem in ET.iterparse(handle, events=("end",)):
                    if row_elem.tag != XLSX_MAIN_NS + "row":
                        continue
                    row_values: list[Any] = []
                    for cell in row_elem.findall(XLSX_MAIN_NS + "c"):
                        ref = cell.attrib.get("r", "")
                        letters = "".join(char for char in ref if char.isalpha())
                        col_index = column_index_from_string(letters) - 1 if letters else len(row_values)
                        cell_type = cell.attrib.get("t")
                        if cell_type == "inlineStr":
                            inline = cell.find(XLSX_MAIN_NS + "is")
                            value = "".join(inline.itertext()) if inline is not None else None
                        else:
                            value_node = cell.find(XLSX_MAIN_NS + "v")
                            raw_value = "".join(value_node.itertext()) if value_node is not None else ""
                            value = raw_value or None
                            if cell_type == "s" and value is not None:
                                value = ("shared", int(value))
                                shared_refs.add(value[1])
                        while len(row_values) <= col_index:
                            row_values.append(None)
                        row_values[col_index] = value
                    rows.append(row_values)
                    row_elem.clear()
                    if len(rows) >= row_limit:
                        break

            shared_strings: dict[int, str] = {}
            if shared_refs and "xl/sharedStrings.xml" in archive.namelist():
                current_index = -1
                with archive.open("xl/sharedStrings.xml") as handle:
                    for _event, item in ET.iterparse(handle, events=("end",)):
                        if item.tag == XLSX_MAIN_NS + "si":
                            current_index += 1
                            if current_index in shared_refs:
                                shared_strings[current_index] = "".join(item.itertext())
                                if len(shared_strings) == len(shared_refs):
                                    item.clear()
                                    break
                            item.clear()

        resolved_rows = [
            [
                shared_strings.get(value[1], "") if isinstance(value, tuple) and value[0] == "shared" else value
                for value in row
            ]
            for row in rows
        ]
        col_count = max((len(row) for row in resolved_rows), default=0)
        return {
            "sheet_names": sheet_names,
            "sheet_name": target_sheet_name,
            "columns": [get_column_letter(index + 1) for index in range(col_count)],
            "rows2d": [
                [(row[index] if index < len(row) else None) for index in range(col_count)]
                for row in resolved_rows
            ],
            "base_row": 0,
            "col_count": col_count,
        }

    def _build_dataframe_from_worksheet_rows(
        self,
        rows: list[list[Any]],
        header_row: int,
        data_start_row: int,
        schema: Any = None,
        date_serial_system: str = "excel_1900",
        date_cleansing: bool = True,
    ) -> pd.DataFrame:
        if not rows:
            return self.attach_dataframe_schema(
                pd.DataFrame(),
                schema_override=schema,
                date_serial_system=date_serial_system,
                date_cleansing=date_cleansing,
            )
        if header_row < 1:
            raise ValueError("header_row は 1 以上で指定してください。")
        if data_start_row < header_row:
            raise ValueError("data_start_row は header_row 以上で指定してください。")
        if header_row > len(rows):
            raise ValueError("header_row が読込範囲を超えています。")
        if data_start_row > len(rows) + 1:
            raise ValueError("data_start_row が読込範囲を超えています。")

        header_cells = rows[header_row - 1]
        headers = [
            str(cell) if cell is not None and str(cell) != "" else f"col_{index}"
            for index, cell in enumerate(header_cells)
        ]

        records: list[dict[str, Any]] = []
        for raw_row in rows[data_start_row - 1:]:
            normalized_row = [self._normalize_excel_value(value) for value in raw_row]
            if any(value is not None and value != "" for value in normalized_row):
                records.append(dict(zip(headers, normalized_row)))
        return self.attach_dataframe_schema(
            pd.DataFrame(records),
            schema_override=schema,
            date_serial_system=date_serial_system,
            date_cleansing=date_cleansing,
        )

    def execute(self, action: str, params: dict[str, Any], context: dict[str, Any]) -> Any:
        if action == "read_excel":
            file_path = params.get('file_path')
            if not file_path:
                raise ValueError("file_path が指定されていません。")
            return self.read_excel(
                path=str(file_path),
                sheet_name=str(params.get('sheet_name', "")),
                header_row=int(params.get('header_row', 1)),
                data_start_row=int(params.get('data_start_row', 2)),
                chunk_size=self._resolve_chunk_size(params.get("chunk_size", DEFAULT_CHUNK_SIZE)),
                schema=params.get('schema'),
                date_cleansing=self._to_bool_flag(params.get("date_cleansing", True)),
            )
        elif action == "read_excel_range":
            file_path = params.get('file_path')
            cell_range = params.get('cell_range')
            if not file_path:
                raise ValueError("file_path が指定されていません。")
            if not cell_range:
                raise ValueError("cell_range が指定されていません。")
            return self.read_excel_range(
                path=str(file_path),
                sheet_name=str(params.get('sheet_name', "")),
                cell_range=str(cell_range),
                header_row=int(params.get('header_row', 1)),
                data_start_row=int(params.get('data_start_row', 2)),
                schema=params.get('schema'),
                date_cleansing=self._to_bool_flag(params.get("date_cleansing", True)),
            )
        elif action == "write_excel":
            input_data = params.get('input_data')
            output_path = self._resolve_output_path(params)
            if not input_data:
                raise ValueError("input_data が指定されていません。")
            if not output_path:
                raise ValueError("output_path が指定されていません。")
            return self.write_excel(
                str(input_data),
                str(output_path),
                str(params.get('sheet_name', 'Sheet1')),
                context,
                str(params.get('mode', 'create_or_replace')),
                params.get('schema'),
            )

    def _resolve_output_path(self, params: dict[str, Any]) -> str | None:
        explicit = params.get("output_path")
        if explicit:
            return str(explicit)

        output_folder = str(params.get("output_folder") or "").strip()
        file_name = str(params.get("file_name") or "").strip()
        if not output_folder or not file_name:
            return None

        if not os.path.splitext(file_name)[1]:
            file_name = f"{file_name}.xlsx"
        return os.path.join(output_folder, file_name)

    # --- 内部ロジック ---

    @staticmethod
    def _resolve_excel_serial_system_from_workbook(workbook) -> str:
        epoch_text = str(getattr(workbook, "epoch", "")).lower()
        if "1904" in epoch_text:
            return "excel_1904"
        if bool(getattr(getattr(workbook, "properties", None), "date1904", False)):
            return "excel_1904"
        return "excel_1900"

    def _resolve_excel_serial_system(self, path: str) -> str:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return self._resolve_excel_serial_system_from_workbook(workbook)
        finally:
            workbook.close()

    def _read_excel_chunked(
        self,
        normalized_path: str,
        sheet_name: str,
        header_row: int,
        data_start_row: int,
        chunk_size: int,
        schema: Any = None,
        date_cleansing: bool = True,
    ) -> pd.DataFrame:
        usecols = self._resolve_usecols_from_schema_origin(schema)
        read_started = time.perf_counter()
        epoch_started = time.perf_counter()
        workbook = load_workbook(normalized_path, read_only=True, data_only=True)
        try:
            try:
                worksheet = workbook[sheet_name] if sheet_name else workbook.active
            except KeyError as exc:
                raise ValueError(f"シートが見つかりませんでした: {sheet_name}") from exc

            date_serial_system = (
                self._resolve_excel_serial_system_from_workbook(workbook)
                if date_cleansing
                else "excel_1900"
            )
            epoch_elapsed_ms = round((time.perf_counter() - epoch_started) * 1000, 1)

            selected_headers: list[str] | None = None
            selected_indices: list[int] = []
            records: list[dict[str, Any]] = []
            chunks: list[pd.DataFrame] = []
            total_rows = 0
            chunk_index = 0

            def flush_chunk() -> None:
                nonlocal records, total_rows, chunk_index
                if not records or selected_headers is None:
                    return
                chunk_index += 1
                chunk = pd.DataFrame.from_records(records, columns=selected_headers)
                chunks.append(chunk)
                total_rows += len(chunk.index)
                elapsed_ms = round((time.perf_counter() - read_started) * 1000, 1)
                self.log_performance("run.connector.chunk.finish", {
                    "connector": "excel_connector",
                    "action": "read_excel",
                    "chunk_index": chunk_index,
                    "chunk_size": chunk_size,
                    "rows": len(chunk.index),
                    "total_rows": total_rows,
                    "elapsed_ms": elapsed_ms,
                })
                self.log_execution(
                    f"Excel chunk read chunk={chunk_index} rows={len(chunk.index)} total_rows={total_rows} elapsed_ms={elapsed_ms}"
                )
                records = []

            for row_number, row_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_number < header_row:
                    continue
                if row_number == header_row:
                    headers = [
                        str(value) if value is not None and str(value) != "" else f"col_{index}"
                        for index, value in enumerate(row_values)
                    ]
                    if usecols:
                        missing = [column for column in usecols if column not in headers]
                        if missing:
                            raise ValueError(
                                f"schema適用エラー(列存在チェック): 指定列が存在しません [{', '.join(missing)}]"
                            )
                        selected_indices = [headers.index(column) for column in usecols]
                    else:
                        selected_indices = list(range(len(headers)))
                    selected_headers = [headers[index] for index in selected_indices]
                    continue
                if row_number < data_start_row:
                    continue
                if selected_headers is None:
                    raise ValueError("header_row が読込範囲を超えています。")

                normalized_values = [
                    self._normalize_excel_value(row_values[index] if index < len(row_values) else None)
                    for index in selected_indices
                ]
                if any(value is not None and value != "" for value in normalized_values):
                    records.append(dict(zip(selected_headers, normalized_values)))
                if len(records) >= chunk_size:
                    flush_chunk()

            if selected_headers is None:
                raise ValueError("header_row が読込範囲を超えています。")
            flush_chunk()
            df = pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame(columns=selected_headers)
        finally:
            workbook.close()

        read_elapsed_ms = round((time.perf_counter() - read_started) * 1000, 1)
        self.log_performance("run.connector.library.finish", {
            "connector": "excel_connector",
            "action": "read_excel",
            "phase": "read_excel",
            "library": "openpyxl/read_only",
            "chunk_size": chunk_size,
            "chunks": chunk_index,
            "elapsed_ms": read_elapsed_ms,
            "rows": len(df.index),
            "cols": len(df.columns),
        })

        if date_cleansing:
            self.log_performance("run.connector.phase.finish", {
                "connector": "excel_connector",
                "action": "read_excel",
                "phase": "epoch_lookup",
                "elapsed_ms": epoch_elapsed_ms,
                "date_serial_system": date_serial_system,
            })
            self.log_execution(f"Excel epoch lookup elapsed_ms={epoch_elapsed_ms} system={date_serial_system}")
        else:
            self.log_performance("run.connector.phase.skipped", {
                "connector": "excel_connector",
                "action": "read_excel",
                "phase": "epoch_lookup",
                "elapsed_ms": epoch_elapsed_ms,
                "reason": "date_cleansing=False",
            })
            self.log_execution("Excel epoch lookup skipped date_cleansing=False")

        schema_started = time.perf_counter()
        result = self.attach_dataframe_schema(
            df,
            schema_override=schema,
            date_serial_system=date_serial_system,
            date_cleansing=date_cleansing,
        )
        schema_elapsed_ms = round((time.perf_counter() - schema_started) * 1000, 1)
        self.log_performance("run.connector.phase.finish", {
            "connector": "excel_connector",
            "action": "read_excel",
            "phase": "schema_apply",
            "elapsed_ms": schema_elapsed_ms,
            "rows": len(result.index),
            "cols": len(result.columns),
        })
        self.log_execution(f"Excel schema apply elapsed_ms={schema_elapsed_ms}")
        self.log_date_parse_metrics(result)
        return result

    def read_excel(
        self,
        path: str,
        sheet_name: str,
        header_row: int,
        data_start_row: int,
        chunk_size: int | None = DEFAULT_CHUNK_SIZE,
        schema: Any = None,
        date_cleansing: bool = True,
    ) -> pd.DataFrame:
        normalized_path = self.normalize_file_path(path)
        if normalized_path is None or not os.path.exists(normalized_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {normalized_path}")
        if header_row < 1:
            raise ValueError("header_row は 1 以上で指定してください。")
        if data_start_row < header_row:
            raise ValueError("data_start_row は header_row 以上で指定してください。")

        resolved_chunk_size = self._resolve_chunk_size(chunk_size)
        return self._read_excel_chunked(
            normalized_path=normalized_path,
            sheet_name=sheet_name,
            header_row=header_row,
            data_start_row=data_start_row,
            chunk_size=resolved_chunk_size,
            schema=schema,
            date_cleansing=date_cleansing,
        )

    def read_excel_range(
        self,
        path: str,
        sheet_name: str,
        cell_range: str,
        header_row: int,
        data_start_row: int,
        schema: Any = None,
        date_cleansing: bool = True,
    ) -> pd.DataFrame:
        normalized_path = self.normalize_file_path(path)
        if normalized_path is None or not os.path.exists(normalized_path):
            raise FileNotFoundError(f"ファイルが見つかりません: {normalized_path}")

        workbook = load_workbook(normalized_path, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
        except KeyError as exc:
            workbook.close()
            raise ValueError(f"シートが見つかりませんでした: {sheet_name}") from exc

        date_serial_system = self._resolve_excel_serial_system_from_workbook(workbook)
        try:
            range_values = worksheet[cell_range]
        except ValueError as exc:
            workbook.close()
            raise ValueError(f"cell_range の指定が不正です: {cell_range}") from exc

        if not isinstance(range_values, tuple):
            range_rows = [[range_values.value]]
        else:
            range_rows = []
            for row in range_values:
                normalized_row = []
                for cell in row:
                    normalized_row.append(cell.value)
                range_rows.append(normalized_row)

        workbook.close()
        result = self._build_dataframe_from_worksheet_rows(
            rows=range_rows,
            header_row=header_row,
            data_start_row=data_start_row,
            schema=schema,
            date_serial_system=date_serial_system,
            date_cleansing=date_cleansing,
        )
        self.log_date_parse_metrics(result)
        return result

    def write_excel(self, input_var: str, output_path: str, sheet_name: str, context: dict[str, Any], mode: str = 'create_or_replace', schema: Any = None):
        """
        mode:
        'create_or_replace': 指定シートを「初期化」して書き込む（他のシートは維持）
        'create_or_insert': 指定シートの「末尾に追記」して書き込む（他のシートは維持）
        """
        normalized_output_path = self.normalize_file_path(output_path)
        if normalized_output_path is None:
            raise ValueError("output_path が指定されていません。")
        mode = str(mode or "create_or_replace").strip()
        if mode == "insert_or_replace":
            mode = "create_or_insert"
        if mode not in {"create_or_replace", "create_or_insert"}:
            raise ValueError(f"未対応の書き込みモードです: {mode}")
        data = context.get(input_var)
        if data is None:
            raise ValueError("データが空です")
        df = self.to_dataframe(data)
        if df.empty:
            raise ValueError("データが空です")
        df_to_write = df.copy()
        if schema is not None and str(schema).strip() != "":
            schema_items = self.parse_schema_definition(schema)
            df_to_write = self.apply_schema_to_dataframe(df_to_write, schema_items)

        if not os.path.exists(normalized_output_path):
            with pd.ExcelWriter(normalized_output_path, engine="openpyxl", mode="w") as writer:
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
            return f"Excel保存完了 [{mode}]: {normalized_output_path} (Sheet: {sheet_name})"

        if mode == 'create_or_replace':
            with pd.ExcelWriter(
                normalized_output_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as writer:
                df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
            return f"Excel保存完了 [{mode}]: {normalized_output_path} (Sheet: {sheet_name})"

        workbook = load_workbook(normalized_output_path)
        try:
            if sheet_name in workbook.sheetnames:
                start_row = workbook[sheet_name].max_row
                write_header = False
            else:
                start_row = 0
                write_header = True
        finally:
            workbook.close()

        with pd.ExcelWriter(
            normalized_output_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay",
        ) as writer:
            df_to_write.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=write_header,
                startrow=start_row,
            )

        return f"Excel保存完了 [{mode}]: {normalized_output_path} (Sheet: {sheet_name})"
