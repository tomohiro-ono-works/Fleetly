import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from connectors.excel_connector import ExcelConnector


class ExcelConnectorTests(unittest.TestCase):
    def _temp_xlsx_path(self):
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        path = Path(handle.name)
        self.addCleanup(lambda: path.unlink(missing_ok=True))
        return path

    def test_read_excel_skips_epoch_lookup_when_date_cleansing_is_false(self):
        connector = ExcelConnector()
        path = self._temp_xlsx_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Orders"
        worksheet.append(["order_date"])
        worksheet.append(["2026-06-14"])
        workbook.save(path)
        workbook.close()
        schema = [{"origin_name": "order_date", "new_name": "order_date", "ziz_datatype": "DATE"}]

        with patch.object(
            connector,
            "_resolve_excel_serial_system_from_workbook",
            side_effect=AssertionError("unexpected epoch lookup"),
        ) as epoch_mock:
            result = connector.read_excel(
                str(path),
                sheet_name="Orders",
                header_row=1,
                data_start_row=2,
                schema=schema,
                date_cleansing=False,
            )

        epoch_mock.assert_not_called()
        self.assertEqual(result.shape, (1, 1))
        self.assertEqual(str(result.iloc[0]["order_date"])[:10], "2026-06-14")

    def test_read_excel_keeps_epoch_lookup_when_date_cleansing_is_true(self):
        connector = ExcelConnector()
        path = self._temp_xlsx_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Orders"
        worksheet.append(["order_date"])
        worksheet.append([44927])
        workbook.save(path)
        workbook.close()
        schema = [{"origin_name": "order_date", "new_name": "order_date", "ziz_datatype": "DATE"}]

        with patch.object(connector, "_resolve_excel_serial_system_from_workbook", return_value="excel_1900") as epoch_mock:
            result = connector.read_excel(
                str(path),
                sheet_name="Orders",
                header_row=1,
                data_start_row=2,
                schema=schema,
                date_cleansing=True,
            )

        epoch_mock.assert_called_once()
        self.assertEqual(result.shape, (1, 1))

    def test_read_excel_rejects_zero_chunk_size(self):
        connector = ExcelConnector()
        path = self._temp_xlsx_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Orders"
        worksheet.append(["order_id"])
        worksheet.append(["order-1"])
        workbook.save(path)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "chunk_size は 1 以上"):
            connector.read_excel(
                str(path),
                sheet_name="Orders",
                header_row=1,
                data_start_row=2,
                chunk_size=0,
            )

    def test_preview_excel_streams_xlsx_without_openpyxl_workbook_load(self):
        connector = ExcelConnector()
        path = self._temp_xlsx_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Orders"
        worksheet.append(["order_id", "amount"])
        for index in range(50):
            worksheet.append([f"order-{index}", index])
        workbook.save(path)
        workbook.close()

        with patch(
            "connectors.excel_connector.load_workbook",
            side_effect=AssertionError("preview must not load workbook"),
        ):
            preview = connector.preview_excel(str(path), sheet_name="Orders", max_rows=30)

        self.assertEqual(preview["sheet_names"], ["Orders"])
        self.assertEqual(preview["sheet_name"], "Orders")
        self.assertEqual(preview["columns"], ["A", "B"])
        self.assertEqual(len(preview["rows2d"]), 30)
        self.assertEqual(preview["rows2d"][0], ["order_id", "amount"])
        self.assertEqual(preview["rows2d"][1], ["order-0", "0"])

    def test_read_excel_chunks_rows_and_logs_progress(self):
        connector = ExcelConnector()
        events = []
        connector.set_execution_logger(None, "step1", lambda event, detail: events.append((event, detail)))
        path = self._temp_xlsx_path()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Orders"
        worksheet.append(["order_id", "amount"])
        for index in range(5):
            worksheet.append([f"order-{index}", index])
        workbook.save(path)
        workbook.close()

        result = connector.read_excel(
            str(path),
            sheet_name="Orders",
            header_row=1,
            data_start_row=2,
            chunk_size=2,
            date_cleansing=False,
        )

        self.assertEqual(result.shape, (5, 2))
        chunk_events = [detail for event, detail in events if event == "run.connector.chunk.finish"]
        self.assertEqual([event["rows"] for event in chunk_events], [2, 2, 1])
        self.assertEqual(chunk_events[-1]["total_rows"], 5)
        connector.clear_execution_logger()


if __name__ == "__main__":
    unittest.main()
