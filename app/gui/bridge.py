import copy
import csv
import json
import logging
import re
import threading
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import yaml
import pandas as pd
from PySide6.QtCore import QObject, Signal, Slot
from openpyxl import load_workbook

from core.type_registry import build_dataframe_schema
from core.sqlbilder_commands import apply_measure_command
from core.workflow_engine import WorkflowEngine
from core.flow_locator import list_flows_local, list_templates_local, register_recent_flow
from core.security_policies import load_security_policies

logger = logging.getLogger("ziz.gui_bridge")


SECRET_FIELD_KEYS = {
    "file_path",
    "folder_path",
    "output_path",
    "directory",
    "output_folder",
    "output_dir",
    "dataset_id",
    "bucket",
}

MODE_EXTENSIONS = {
    "workflow": ".zizw",
    "dataflow": ".zizd",
    "query-builder": ".zizq",
}


def _iso_now():
    return datetime.now(timezone.utc).isoformat()


def _safe_text(value):
    return str(value or "").strip()


def _is_hidden_ref(value):
    return isinstance(value, str) and re.fullmatch(r"\{\{hidden\.[^}]+\}\}", value.strip()) is not None


def _sanitize_hidden_scope(step_name):
    text = re.sub(r"[^a-zA-Z0-9_]+", "_", _safe_text(step_name))
    return text or "global"


class BridgeApiError(Exception):
    def __init__(self, code, message):
        super().__init__(str(message or ""))
        self.code = str(code or "E_INTERNAL")
        self.message = str(message or "")


class _RunLogHandler(logging.Handler):
    def __init__(self, runtime, run_id):
        super().__init__(level=logging.INFO)
        self._runtime = runtime
        self._run_id = run_id

    def emit(self, record):
        try:
            message = self._runtime.mask_log_message(self._run_id, record.getMessage())
            self._runtime.emit_event("run.log", {
                "run_id": self._run_id,
                "ts": datetime.fromtimestamp(record.created, timezone.utc).isoformat(),
                "level": record.levelname,
                "message": message,
            })
        except Exception:
            # ログ送信失敗で実行全体を落とさない
            return


class BridgeRuntime:
    PROTOCOL_VERSION = "1.0"

    def __init__(
        self,
        base_dir,
        *,
        debug=False,
        pick_file_callback=None,
        pick_folder_callback=None,
        edit_file_callback=None,
        edit_folder_callback=None,
        open_flow_callback=None,
        save_flow_callback=None,
        window_control_callback=None,
    ):
        self.base_dir = Path(base_dir).resolve()
        self.debug = bool(debug)
        self.pick_file_callback = pick_file_callback
        self.pick_folder_callback = pick_folder_callback
        self.edit_file_callback = edit_file_callback
        self.edit_folder_callback = edit_folder_callback
        self.open_flow_callback = open_flow_callback
        self.save_flow_callback = save_flow_callback
        self.window_control_callback = window_control_callback
        self._flow_tokens = {}
        self._event_sink = None
        self._lock = threading.RLock()

        self.current_flow_path = None
        self.current_file_name = ""
        self.current_mode = ""
        self.hidden_values = {}
        self.hidden_meta = {}
        self._hidden_counters = {}
        self.runs = {}
        self.latest_by_flow = {}
        self.active_run_by_flow = {}
        self._unsaved_flow_uuid = uuid.uuid4().hex
        self._execution_log_path = (self.base_dir / "logs" / "execution.log").resolve()
        self._execution_log_path.parent.mkdir(parents=True, exist_ok=True)

    def set_event_sink(self, callback):
        self._event_sink = callback

    def emit_event(self, message_type, payload):
        if not self._event_sink:
            return
        message = {
            "v": self.PROTOCOL_VERSION,
            "kind": "evt",
            "type": message_type,
            "ts": _iso_now(),
            "payload": payload or {},
        }
        self._event_sink(message)

    def handle_message(self, raw_text):
        try:
            message = json.loads(str(raw_text or ""))
        except Exception:
            return self._error_response(None, "unknown", "E_VALIDATION", "メッセージが JSON ではありません。")

        message_id = str(message.get("id") or "")
        message_type = str(message.get("type") or "")
        version = str(message.get("v") or "")
        kind = str(message.get("kind") or "")
        payload = message.get("payload") or {}

        if version != self.PROTOCOL_VERSION:
            return self._error_response(message_id, message_type, "E_CONTRACT_VERSION_MISMATCH", "プロトコルバージョンが一致しません。")
        if kind != "cmd":
            return self._error_response(message_id, message_type, "E_VALIDATION", "cmd メッセージのみ受け付けます。")

        try:
            if message_type == "app.getStatus":
                return self._success_response(message_id, message_type, self._handle_app_get_status())
            if message_type == "app.logUiEvent":
                return self._success_response(message_id, message_type, self._handle_app_log_ui_event(payload))
            if message_type == "app.windowControl":
                return self._success_response(message_id, message_type, self._handle_app_window_control(payload))
            if message_type == "flow.list":
                return self._success_response(message_id, message_type, self._handle_flow_list(payload))
            if message_type == "flow.load":
                return self._success_response(message_id, message_type, self._handle_flow_load(payload))
            if message_type == "flow.save":
                return self._success_response(message_id, message_type, self._handle_flow_save(payload))
            if message_type == "flow.run":
                return self._success_response(message_id, message_type, self._handle_flow_run(payload))
            if message_type == "run.cancel":
                return self._success_response(message_id, message_type, self._handle_run_cancel(payload))
            if message_type == "result.getSummary":
                return self._success_response(message_id, message_type, self._handle_result_get_summary(payload))
            if message_type == "result.getSchema":
                return self._success_response(message_id, message_type, self._handle_result_get_schema(payload))
            if message_type == "result.getPreview":
                return self._success_response(message_id, message_type, self._handle_result_get_preview(payload))
            if message_type == "result.getDatavolume":
                return self._success_response(message_id, message_type, self._handle_result_get_datavolume(payload))
            if message_type == "file.pickFile":
                return self._success_response(message_id, message_type, self._handle_file_pick_file(payload))
            if message_type == "file.pickFolder":
                return self._success_response(message_id, message_type, self._handle_file_pick_folder(payload))
            if message_type == "preview.readExcel":
                return self._success_response(message_id, message_type, self._handle_preview_read_excel(payload))
            if message_type == "preview.readCsv":
                return self._success_response(message_id, message_type, self._handle_preview_read_csv(payload))
            if message_type == "sqlbilder.applyMeasure":
                return self._success_response(message_id, message_type, self._handle_sqlbilder_apply_measure(payload))
        except ValueError as error:
            return self._error_response(message_id, message_type, "E_VALIDATION", str(error))
        except BridgeApiError as error:
            return self._error_response(message_id, message_type, error.code, error.message)
        except FileNotFoundError as error:
            return self._error_response(message_id, message_type, "E_NOT_FOUND", str(error))
        except PermissionError as error:
            return self._error_response(message_id, message_type, "E_ACCESS_DENIED", str(error))
        except Exception:
            return self._error_response(message_id, message_type, "E_INTERNAL", "内部エラーが発生しました。")

        return self._error_response(message_id, message_type, "E_ACCESS_DENIED", "未許可の API です。")

    def _handle_app_get_status(self):
        policies = load_security_policies(self.base_dir)
        return {
            "app": "zizai",
            "host": "pyside6-qtwebengine",
            "protocol_version": self.PROTOCOL_VERSION,
            "gui_mode": "webview",
            "security": {
                "external_requests_blocked": True,
                "navigation_locked": True,
                "devtools_context_menu_disabled": not self.debug,
                "devtools_enabled": self.debug,
                "remote_debugging_disabled": True,
            },
            "capabilities": [
                "app.getStatus",
                "app.logUiEvent",
                "app.windowControl",
                "flow.list",
                "flow.load",
                "flow.save",
                "flow.run",
                "run.cancel",
                "result.getSummary",
                "result.getSchema",
                "result.getPreview",
                "result.getDatavolume",
                "file.pickFile",
                "file.pickFolder",
                "preview.readExcel",
                "preview.readCsv",
                "sqlbilder.applyMeasure",
            ],
            "security_policies": {
                "loaded": bool(policies.get("loaded")),
                "path": str(policies.get("path") or ""),
                "api_profile_count": len(policies.get("apis", {}).get("profiles", {})),
                "web_allowlist_count": len(policies.get("web", {}).get("allowlist", [])),
            },
        }

    def _handle_app_log_ui_event(self, payload):
        action = _safe_text((payload or {}).get("action")) or "unknown"
        source = _safe_text((payload or {}).get("source")) or "ui"
        elapsed_ms = (payload or {}).get("elapsed_ms")
        detail = payload.get("detail") if isinstance(payload, dict) else None
        detail_suffix = ""
        if isinstance(detail, dict) and detail:
            parts = []
            for key, value in detail.items():
                text = _safe_text(value)
                if text:
                    parts.append(f"{key}={text}")
            if parts:
                detail_suffix = " " + " ".join(parts)
        elapsed_suffix = f" elapsed_ms={elapsed_ms}" if elapsed_ms is not None else ""
        logger.info("[ui-event] source=%s action=%s%s%s", source, action, elapsed_suffix, detail_suffix)
        return {"logged": True}

    def _handle_app_window_control(self, payload):
        action = _safe_text((payload or {}).get("action"))
        if action not in {"minimize", "maximize", "close", "drag"}:
            raise ValueError("action が不正です。")
        if not callable(self.window_control_callback):
            raise ValueError("window control は利用できません。")
        state = self.window_control_callback(action)
        return {
            "accepted": True,
            "action": action,
            "state": str(state or ""),
        }

    def _handle_flow_list(self, payload):
        started = time.perf_counter()
        scope = str((payload or {}).get("scope") or "local")
        kind = str((payload or {}).get("kind") or "recent")
        if scope not in {"local", "workspace"}:
            raise ValueError("scope が不正です。")
        if kind not in {"recent", "template"}:
            raise ValueError("kind が不正です。")

        items = []
        source_items = list_templates_local() if kind == "template" else list_flows_local()
        for item in source_items:
            path = str(item.get("path") or "")
            token = self._register_flow_token(path)
            directory = str(item.get("directory") or "")
            items.append({
                "flow_token": token,
                "display_name": str(item.get("filename") or ""),
                "display_hint": directory,
                "modified_at": datetime.fromtimestamp(float(item.get("modified_at") or 0), tz=timezone.utc).isoformat(),
            })
        response = {
            "scope": scope,
            "kind": kind,
            "items": items,
        }
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        logger.info("[bridge] flow.list kind=%s count=%s elapsed_ms=%s", kind, len(items), elapsed_ms)
        return response

    def _handle_flow_load(self, payload):
        started = time.perf_counter()
        ref = _safe_text((payload or {}).get("ref") or (payload or {}).get("flow_token"))
        if ref:
            flow_path = self._flow_tokens.get(ref)
        else:
            flow_path = self.open_flow_callback() if self.open_flow_callback else None

        if not flow_path:
            response = {
                "selected": False,
            }
            elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
            logger.info("[bridge] flow.load cancelled elapsed_ms=%s", elapsed_ms)
            return response

        resolved_path = Path(flow_path).resolve()
        if not resolved_path.exists():
            raise FileNotFoundError(f"フローが見つかりません: {resolved_path}")

        with resolved_path.open("r", encoding="utf-8") as handle:
            config = yaml.safe_load(handle) or {}
        if not isinstance(config, dict):
            raise ValueError("フロー定義は辞書形式である必要があります。")

        with self._lock:
            self.current_flow_path = str(resolved_path)
            self.current_file_name = resolved_path.name
            self.current_mode = self._resolve_mode(config, resolved_path.name)
            self.hidden_values = {}
            self.hidden_meta = {}
            self._hidden_counters = {}
            web_flow = self._hide_sensitive_values(config)
        register_recent_flow(str(resolved_path), opened_at_iso=_iso_now())

        response = {
            "selected": True,
            "mode": self.current_mode,
            "file_name": self.current_file_name,
            "flow": web_flow,
            "hidden_bindings": copy.deepcopy(self.hidden_meta),
        }
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        logger.info("[bridge] flow.load file=%s mode=%s hidden=%s elapsed_ms=%s", self.current_file_name, self.current_mode, len(self.hidden_values), elapsed_ms)
        return response

    def _handle_flow_save(self, payload):
        started = time.perf_counter()
        mode = _safe_text((payload or {}).get("mode")) or self.current_mode or "workflow"
        flow = copy.deepcopy((payload or {}).get("flow") or {})
        if not isinstance(flow, dict):
            raise ValueError("flow はオブジェクトで指定してください。")
        previous_flow_key = self._build_flow_key(mode, self.current_flow_path)

        restored_flow = self._restore_hidden_values(flow)
        suggested_name = _safe_text((payload or {}).get("file_name")) or self.current_file_name or self._build_default_flow_file_name(mode)
        current_path = self.current_flow_path
        logger.info(
            "[bridge] flow.save begin mode=%s suggested_name=%s current_flow_path=%s",
            mode,
            suggested_name,
            current_path or "",
        )
        if self.save_flow_callback:
            target_path = self.save_flow_callback(mode, suggested_name, current_path)
        else:
            target_path = current_path
        if not target_path:
            response = {
                "saved": False,
                "file_name": suggested_name,
            }
            elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
            logger.info("[bridge] flow.save cancelled mode=%s elapsed_ms=%s", mode, elapsed_ms)
            return response

        resolved_path = Path(target_path).resolve()
        resolved_path.parent.mkdir(parents=True, exist_ok=True)
        with resolved_path.open("w", encoding="utf-8") as handle:
            yaml.safe_dump(restored_flow, handle, allow_unicode=True, sort_keys=False)

        with self._lock:
            self.current_flow_path = str(resolved_path)
            self.current_file_name = resolved_path.name
            self.current_mode = mode
        saved_flow_key = self._build_flow_key(mode, str(resolved_path))
        self._migrate_flow_state(previous_flow_key, saved_flow_key)
        register_recent_flow(str(resolved_path), opened_at_iso=_iso_now())

        response = {
            "saved": True,
            "file_name": resolved_path.name,
        }
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        logger.info("[bridge] flow.save file=%s mode=%s elapsed_ms=%s", resolved_path.name, mode, elapsed_ms)
        return response

    def _handle_flow_run(self, payload):
        started = time.perf_counter()
        mode = _safe_text((payload or {}).get("mode")) or self.current_mode or "workflow"
        flow = copy.deepcopy((payload or {}).get("flow") or {})
        if not isinstance(flow, dict):
            raise ValueError("flow はオブジェクトで指定してください。")

        requested_step_id = _safe_text((payload or {}).get("step_id"))
        resolved_flow = self._restore_hidden_values(flow)
        run_id = f"run_{uuid.uuid4().hex[:12]}"
        trace_id = f"trace_{uuid.uuid4().hex[:12]}"
        started_at = _iso_now()
        cancel_event = threading.Event()
        flow_name = _safe_text(((resolved_flow.get("metadata") or {}).get("name"))) or "Untitled"
        secret_values = self._collect_secret_values(resolved_flow)
        flow_key = self._build_flow_key(mode, self.current_flow_path)
        seed_context = self._get_latest_flow_context(flow_key) if requested_step_id else {}
        if self.current_flow_path:
            secret_values.add(self.current_flow_path)

        session = {
            "run_id": run_id,
            "trace_id": trace_id,
            "status": "running",
            "started_at": started_at,
            "finished_at": None,
            "flow_name": flow_name,
            "mode": mode,
            "flow_path": self.current_flow_path or "<unsaved>",
            "step_count": 1 if requested_step_id else len((resolved_flow.get("steps") or []) if isinstance(resolved_flow.get("steps"), list) else []),
            "report": None,
            "cancel_event": cancel_event,
            "error": None,
            "secret_values": secret_values,
            "requested_step_id": requested_step_id,
            "seed_context": seed_context,
            "flow_key": flow_key,
        }
        with self._lock:
            active_run_id = _safe_text(self.active_run_by_flow.get(flow_key))
            active_session = self.runs.get(active_run_id) if active_run_id else None
            is_conflict = bool(active_session and str(active_session.get("status") or "") == "running")
            if is_conflict:
                raise BridgeApiError("E_CONFLICT", "同じフローが実行中です。完了またはキャンセル後に再実行してください。")
            if active_run_id:
                self.active_run_by_flow.pop(flow_key, None)
            self.runs[run_id] = session
            self.active_run_by_flow[flow_key] = run_id
            self.current_mode = mode

        self._append_execution_log("run.start", {
            "run_id": run_id,
            "flow_key": flow_key,
            "mode": mode,
            "step_id": requested_step_id or "",
            "flow_name": flow_name,
        })

        worker = threading.Thread(
            target=self._run_flow_worker,
            args=(run_id, resolved_flow),
            daemon=True,
        )
        session["thread"] = worker
        worker.start()

        response = {
            "run_id": run_id,
            "accepted": True,
        }
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        logger.info("[bridge] flow.run accepted run_id=%s mode=%s step_id=%s elapsed_ms=%s", run_id, mode, requested_step_id or "-", elapsed_ms)
        return response

    def _handle_run_cancel(self, payload):
        run_id = _safe_text((payload or {}).get("run_id"))
        if not run_id:
            raise ValueError("run_id は必須です。")
        with self._lock:
            session = self.runs.get(run_id)
        if not session:
            raise FileNotFoundError("対象の実行が見つかりません。")
        session["cancel_event"].set()
        self.emit_event("run.progress", {
            "run_id": run_id,
            "stage": "cancel_requested",
            "percent": None,
            "message": "キャンセル要求を受け付けました。",
        })
        return {
            "run_id": run_id,
            "accepted": True,
        }

    def _handle_result_get_summary(self, payload):
        session = self._require_run_session((payload or {}).get("run_id"))
        report = session.get("report") or {}
        steps = report.get("steps") or []
        return {
            "flow_name": session.get("flow_name") or report.get("flow_name") or "Untitled",
            "status": session.get("status") or report.get("status") or "error",
            "started_at": session.get("started_at"),
            "finished_at": session.get("finished_at"),
            "duration_ms": self._compute_duration_ms(session.get("started_at"), session.get("finished_at")),
            "step_count": session.get("step_count") or len(steps),
            "success_count": sum(1 for step in steps if step.get("status") == "success"),
            "error_count": sum(1 for step in steps if step.get("status") == "error"),
            "run_id": session.get("run_id"),
            "trace_id": session.get("trace_id"),
            "error": session.get("error"),
        }

    def _handle_result_get_schema(self, payload):
        dataframe = self._require_latest_step_result(payload)
        if not hasattr(dataframe, "columns") or not hasattr(dataframe, "attrs"):
            raise ValueError("指定ステップの結果は表データではありません。")

        existing_schema = dataframe.attrs.get("ziz_schema")
        if isinstance(existing_schema, list) and existing_schema:
            schema_items = existing_schema
        else:
            schema_items = build_dataframe_schema(dataframe)
        return {
            "columns": [
                {
                    "origin_name": str(item.get("origin_name") or item.get("name_ja") or item.get("name_en") or ""),
                    "new_name": str(item.get("new_name") or item.get("name_en") or item.get("origin_name") or ""),
                    "description": str(item.get("description") or item.get("name_ja") or item.get("origin_name") or ""),
                    "ziz_datatype": str(item.get("ziz_datatype") or ""),
                }
                for item in schema_items
            ]
        }

    def _handle_result_get_preview(self, payload):
        dataframe = self._require_latest_step_result(payload)
        if not hasattr(dataframe, "columns") or not hasattr(dataframe, "head"):
            raise ValueError("指定ステップの結果は表データではありません。")

        preview = dataframe.head(100).copy()
        columns = [str(column) for column in preview.columns]
        rows = []
        for _, row in preview.iterrows():
            values = []
            for value in row.tolist():
                if self._is_missing_preview_value(value):
                    values.append("")
                    continue
                values.append(str(value))
            rows.append(values)

        return {
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "truncated": bool(len(dataframe.index) > len(rows)),
        }

    def _handle_result_get_datavolume(self, payload):
        dataframe = self._require_latest_step_result(payload)
        if not hasattr(dataframe, "columns") or not hasattr(dataframe, "index"):
            raise ValueError("指定ステップの結果は表データではありません。")

        total_rows = len(dataframe.index)
        top_n = int((payload or {}).get("top_n") or 5)
        top_n = max(1, min(top_n, 20))
        columns = []
        for column in dataframe.columns:
            series = dataframe[column]
            normalized = self._normalize_datavolume_series(series)
            counts = normalized.value_counts(dropna=False).head(top_n)
            items = []
            for item_value, item_count in counts.items():
                ratio = (int(item_count) / total_rows) * 100 if total_rows else 0
                items.append({
                    "value": str(item_value),
                    "count": int(item_count),
                    "ratio": round(ratio, 1),
                })
            columns.append({
                "name": str(column),
                "dtype": str(series.dtype),
                "items": items,
            })

        return {
            "row_count": total_rows,
            "columns": columns,
            "top_n": top_n,
        }

    def _is_missing_preview_value(self, value):
        if value is None:
            return True
        try:
            result = pd.isna(value)
        except Exception:
            return False
        if isinstance(result, bool):
            return result
        return False

    def _normalize_datavolume_series(self, series):
        object_series = series.astype(object)
        return object_series.where(pd.notna(object_series), "NULL").astype(str)

    def _handle_file_pick_file(self, payload):
        if not self.pick_file_callback and not self.edit_file_callback:
            raise RuntimeError("ファイル選択ダイアログが利用できません。")
        title = _safe_text((payload or {}).get("title")) or "ファイルを選択"
        filters = (payload or {}).get("filters") or []
        current_ref = _safe_text((payload or {}).get("current_ref")) or None
        current_value = self._resolve_picker_initial_value(payload, current_ref)
        if self.edit_file_callback:
            selected_path = self.edit_file_callback(title, current_value, filters)
        else:
            selected_path = self.pick_file_callback(title, filters)
        if not selected_path:
            return {"selected": False}
        ref, meta = self._store_hidden_value(
            step_name=_safe_text((payload or {}).get("step_name")) or "global",
            field_key=_safe_text((payload or {}).get("field_key")) or "file_path",
            actual_value=str(selected_path),
            current_ref=current_ref,
        )
        return {
            "ref": ref,
            "display_name": meta.get("display_name") or "",
            "display_hint": meta.get("display_hint") or "",
            "selected": True,
        }

    def _handle_file_pick_folder(self, payload):
        if not self.pick_folder_callback and not self.edit_folder_callback:
            raise RuntimeError("フォルダ選択ダイアログが利用できません。")
        title = _safe_text((payload or {}).get("title")) or "フォルダを選択"
        current_ref = _safe_text((payload or {}).get("current_ref")) or None
        current_value = self._resolve_picker_initial_value(payload, current_ref)
        if self.edit_folder_callback:
            selected_path = self.edit_folder_callback(title, current_value)
        else:
            selected_path = self.pick_folder_callback(title)
        if not selected_path:
            return {"selected": False}
        ref, meta = self._store_hidden_value(
            step_name=_safe_text((payload or {}).get("step_name")) or "global",
            field_key=_safe_text((payload or {}).get("field_key")) or "folder_path",
            actual_value=str(selected_path),
            current_ref=current_ref,
        )
        return {
            "ref": ref,
            "display_name": meta.get("display_name") or "",
            "display_hint": meta.get("display_hint") or "",
            "selected": True,
        }

    def _handle_preview_read_excel(self, payload):
        actual_path, ref, meta = self._resolve_hidden_or_current_path(
            payload,
            field_key=_safe_text((payload or {}).get("field_key")) or "file_path",
        )
        if not actual_path:
            raise ValueError("Excel ファイルが選択されていません。")

        selected_sheet = _safe_text((payload or {}).get("sheet_name"))
        workbook = load_workbook(actual_path, read_only=True, data_only=True)
        sheet_names = list(workbook.sheetnames or [])
        if not sheet_names:
            raise ValueError("シートがありません。")
        target_sheet = selected_sheet if selected_sheet in sheet_names else sheet_names[0]
        worksheet = workbook[target_sheet]
        rows2d = self._load_excel_top_rows(worksheet, max_rows=30)
        col_count = max((len(row) for row in rows2d), default=0)
        columns = [self._col_index_to_letters(index) for index in range(col_count)]
        fixed_rows = [self._pad_row(row, col_count) for row in rows2d]
        workbook.close()
        return {
            "ref": ref,
            "display_name": meta.get("display_name") or Path(actual_path).name,
            "display_hint": meta.get("display_hint") or "",
            "file_name": Path(actual_path).name,
            "sheet_names": sheet_names,
            "sheet_name": target_sheet,
            "columns": columns,
            "rows2d": fixed_rows,
            "base_row": 0,
            "col_count": col_count,
        }

    def _handle_preview_read_csv(self, payload):
        actual_path, ref, meta = self._resolve_hidden_or_current_path(
            payload,
            field_key=_safe_text((payload or {}).get("field_key")) or "file_path",
        )
        if not actual_path:
            raise ValueError("CSV ファイルが選択されていません。")

        encoding = self._normalize_preview_encoding((payload or {}).get("encoding"))
        delimiter = self._normalize_preview_delimiter((payload or {}).get("delimiter"))
        rows = self._load_csv_rows(actual_path, encoding, delimiter, max_rows=100)
        preview_rows = rows[:30]
        col_count = max((len(row) for row in rows), default=0)
        columns = [self._col_index_to_letters(index) for index in range(col_count)]
        return {
            "ref": ref,
            "display_name": meta.get("display_name") or Path(actual_path).name,
            "display_hint": meta.get("display_hint") or "",
            "file_name": Path(actual_path).name,
            "encoding": encoding,
            "delimiter": delimiter,
            "columns": columns,
            "rows2d": [self._pad_row(row, col_count) for row in preview_rows],
            "schema_rows2d": [self._pad_row(row, col_count) for row in rows],
            "base_row": 0,
            "col_count": col_count,
        }

    def _handle_sqlbilder_apply_measure(self, payload):
        measure_type = _safe_text((payload or {}).get("measure_type"))
        sql_text = str((payload or {}).get("sql_text") or "")
        selection_start_line = int((payload or {}).get("selection_start_line") or 0)
        selection_end_line = int((payload or {}).get("selection_end_line") or selection_start_line)
        return {
            "replacement": apply_measure_command(
                sql_text,
                measure_type=measure_type,
                selection_start_line=selection_start_line,
                selection_end_line=selection_end_line,
            )
        }

    def _resolve_picker_initial_value(self, payload, current_ref):
        if current_ref and current_ref in self.hidden_values:
            return str(self.hidden_values.get(current_ref) or "")
        return _safe_text((payload or {}).get("current_value"))

    def _resolve_hidden_or_current_path(self, payload, *, field_key):
        current_ref = _safe_text((payload or {}).get("current_ref")) or None
        current_value = self._resolve_picker_initial_value(payload, current_ref)
        actual_path = str(current_value or "")
        if not actual_path:
            raise ValueError("対象パスが未設定です。")
        path_obj = Path(actual_path).resolve()
        if not path_obj.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {path_obj}")
        meta = self.hidden_meta.get(current_ref) if current_ref and current_ref in self.hidden_meta else self._build_hidden_meta(field_key, actual_path)
        return str(path_obj), current_ref or "", meta

    def _load_excel_top_rows(self, worksheet, *, max_rows):
        rows = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if row_index >= max_rows:
                break
            normalized = [self._normalize_excel_preview_value(cell) for cell in list(row)]
            rows.append(normalized)
        return rows

    def _normalize_excel_preview_value(self, value):
        if value is None:
            return None
        if pd.isna(value):
            return None
        if hasattr(value, "isoformat"):
            try:
                text = value.isoformat()
                return text[:10] if text.endswith("T00:00:00") else text
            except Exception:
                return str(value)
        return value

    def _load_csv_rows(self, actual_path, encoding, delimiter, *, max_rows):
        normalized_encoding = self._normalize_preview_encoding(encoding)
        normalized_delimiter = self._normalize_preview_delimiter(delimiter)
        rows = []
        with open(actual_path, "r", encoding=normalized_encoding, newline="") as handle:
            reader = csv.reader(handle, delimiter=normalized_delimiter)
            for row_index, row in enumerate(reader):
                if row_index >= max_rows:
                    break
                rows.append([cell if cell != "" else None for cell in row])
        return rows

    def _normalize_preview_encoding(self, value):
        raw = _safe_text(value).lower() or "utf-8"
        if raw == "utf8":
            return "utf-8"
        if raw == "cp932":
            return "cp932"
        return raw

    def _normalize_preview_delimiter(self, value):
        raw = str(value or ",")
        if raw == "\\t" or raw.lower() == "tab":
            return "\t"
        return raw

    def _col_index_to_letters(self, idx0):
        n = int(idx0) + 1
        text = ""
        while n > 0:
            r = (n - 1) % 26
            text = chr(65 + r) + text
            n = (n - 1) // 26
        return text

    def _pad_row(self, row, col_count):
        return [(row[index] if index < len(row) else None) for index in range(col_count)]

    def _run_flow_worker(self, run_id, resolved_flow):
        with self._lock:
            session = self.runs.get(run_id)
        if not session:
            return

        logger = self._create_run_logger(run_id)
        log_handler = _RunLogHandler(self, run_id)
        logger.addHandler(log_handler)
        engine = WorkflowEngine(logger, step_status_callback=lambda detail: self._emit_step_status(run_id, detail))
        self.emit_event("run.progress", {
            "run_id": run_id,
            "stage": "running",
            "percent": None,
            "message": "実行を開始しました。",
        })

        try:
            report = engine.run_flow_from_config(
                resolved_flow,
                flow_path=session.get("flow_path") or "<gui>",
                cancel_event=session.get("cancel_event"),
                initial_context=session.get("seed_context") or {},
                only_step_id=session.get("requested_step_id") or None,
            )
            session["report"] = {
                "status": str(report.get("status") or ""),
                "error": str(report.get("error") or ""),
                "steps": [
                    {
                        "step_id": _safe_text(step.get("step_id")),
                        "status": _safe_text(step.get("status")),
                    }
                    for step in (report.get("steps") or [])
                ],
                "flow_name": _safe_text(report.get("flow_name") or session.get("flow_name") or "Untitled"),
            }
            session["finished_at"] = _iso_now()
            self._update_latest_by_flow(
                session.get("flow_key"),
                session.get("seed_context"),
                report,
            )
            self._append_execution_log("run.finish", {
                "run_id": run_id,
                "flow_key": session.get("flow_key") or "",
                "status": str(report.get("status") or ""),
                "step_count": len(report.get("steps") or []),
                "error": self._normalize_log_error(report.get("error")),
            })
            for step in (report.get("steps") or []):
                self._append_execution_log("run.step", {
                    "run_id": run_id,
                    "flow_key": session.get("flow_key") or "",
                    "step_id": _safe_text(step.get("step_id")),
                    "status": _safe_text(step.get("status")),
                    "error": self._normalize_log_error(step.get("error")),
                })

            if report.get("status") == "success":
                session["status"] = "success"
                self.emit_event("run.completed", {
                    "run_id": run_id,
                    "status": "success",
                    "trace_id": session.get("trace_id"),
                })
            elif report.get("status") == "cancelled" or report.get("cancelled"):
                session["status"] = "cancelled"
                session["error"] = {
                    "code": "E_CANCELLED",
                    "message": "実行がキャンセルされました。",
                    "detail": {},
                    "retryable": False,
                    "trace_id": session.get("trace_id"),
                }
                self.emit_event("run.failed", {
                    "run_id": run_id,
                    "status": "cancelled",
                    "trace_id": session.get("trace_id"),
                })
            else:
                session["status"] = "error"
                session["error"] = {
                    "code": "E_INTERNAL",
                    "message": str(report.get("error") or "実行に失敗しました。"),
                    "detail": {},
                    "retryable": False,
                    "trace_id": session.get("trace_id"),
                }
                self.emit_event("run.failed", {
                    "run_id": run_id,
                    "status": "error",
                    "trace_id": session.get("trace_id"),
                })
        except Exception:
            session["finished_at"] = _iso_now()
            session["status"] = "error"
            session["report"] = {
                "status": "error",
                "error": "内部エラーが発生しました。",
                "steps": [],
                "flow_name": _safe_text(session.get("flow_name") or "Untitled"),
            }
            session["error"] = {
                "code": "E_INTERNAL",
                "message": "内部エラーが発生しました。",
                "detail": {},
                "retryable": False,
                "trace_id": session.get("trace_id"),
            }
            self._append_execution_log("run.finish", {
                "run_id": run_id,
                "flow_key": session.get("flow_key") or "",
                "status": "error",
                "step_count": 0,
                "error": "内部エラーが発生しました。",
            })
            self.emit_event("run.failed", {
                "run_id": run_id,
                "status": "error",
                "trace_id": session.get("trace_id"),
            })
        finally:
            flow_key = _safe_text(session.get("flow_key"))
            with self._lock:
                if self.active_run_by_flow.get(flow_key) == run_id:
                    self.active_run_by_flow.pop(flow_key, None)
            logger.removeHandler(log_handler)
            log_handler.close()

    def _emit_step_status(self, run_id, detail):
        payload = {
            "run_id": run_id,
            "step_id": str(detail.get("step_id") or ""),
            "status": str(detail.get("status") or ""),
            "message": str(detail.get("message") or ""),
        }
        self.emit_event("run.stepStatus", payload)

    def _create_run_logger(self, run_id):
        logger = logging.getLogger(f"ziz.gui.{run_id}")
        logger.setLevel(logging.INFO)
        logger.handlers = []
        logger.propagate = True
        return logger

    def _compute_duration_ms(self, started_at, finished_at):
        if not started_at or not finished_at:
            return None
        try:
            start_dt = datetime.fromisoformat(str(started_at))
            end_dt = datetime.fromisoformat(str(finished_at))
        except ValueError:
            return None
        return max(0, int((end_dt - start_dt).total_seconds() * 1000))

    def _require_run_session(self, run_id):
        key = _safe_text(run_id)
        if not key:
            raise ValueError("run_id は必須です。")
        with self._lock:
            session = self.runs.get(key)
        if not session:
            raise FileNotFoundError("対象の実行が見つかりません。")
        return session

    def _build_flow_key(self, mode, flow_path=None):
        mode_text = _safe_text(mode) or "workflow"
        path_text = _safe_text(flow_path)
        if path_text:
            try:
                normalized_path = str(Path(path_text).resolve())
            except Exception:
                normalized_path = path_text
        else:
            normalized_path = f"<unsaved:{self._unsaved_flow_uuid}>"
        return f"{mode_text}|{normalized_path}"

    def _get_latest_flow_context(self, flow_key):
        key = _safe_text(flow_key)
        if not key:
            return {}
        with self._lock:
            latest = self.latest_by_flow.get(key) or {}
            context = latest.get("context") if isinstance(latest, dict) else {}
        return copy.copy(context) if isinstance(context, dict) else {}

    def _update_latest_by_flow(self, flow_key, seed_context, report):
        key = _safe_text(flow_key)
        if not key:
            return
        with self._lock:
            existing = self.latest_by_flow.get(key)
            if not isinstance(existing, dict):
                existing = {}
            current_context = copy.copy(existing.get("context")) if isinstance(existing.get("context"), dict) else {}
            context = copy.copy(seed_context) if isinstance(seed_context, dict) else current_context
            step_data = dict(existing.get("step_data") or {})
            step_status = dict(existing.get("step_status") or {})
            for step in (report.get("steps") or []):
                step_id = _safe_text(step.get("step_id"))
                if not step_id:
                    continue
                status = _safe_text(step.get("status")) or "unknown"
                step_status[step_id] = status
                if status != "success":
                    continue
                result = step.get("result")
                step_data[step_id] = result
                output_var = _safe_text(step.get("output_variable")) or step_id
                context[output_var] = result
            self.latest_by_flow[key] = {
                "context": context,
                "step_data": step_data,
                "step_status": step_status,
            }

    def _require_latest_step_result(self, payload):
        step_id = _safe_text((payload or {}).get("step_id"))
        if not step_id:
            raise ValueError("step_id は必須です。")
        mode = _safe_text((payload or {}).get("mode")) or self.current_mode or "workflow"
        flow_key = _safe_text((payload or {}).get("flow_key")) or self._build_flow_key(mode, self.current_flow_path)
        with self._lock:
            latest = self.latest_by_flow.get(flow_key)
            step_data = (latest or {}).get("step_data") if isinstance(latest, dict) else {}
            step_status = (latest or {}).get("step_status") if isinstance(latest, dict) else {}
            if not isinstance(step_data, dict):
                step_data = {}
            if not isinstance(step_status, dict):
                step_status = {}
            has_data = step_id in step_data
            status = _safe_text(step_status.get(step_id))
            result = step_data.get(step_id)
        if not has_data:
            if status == "error":
                raise FileNotFoundError("指定ステップの最新実行は失敗しており、成功データがありません。")
            raise FileNotFoundError("対象のステップ結果が見つかりません。")
        return result

    def _migrate_flow_state(self, old_flow_key, new_flow_key):
        old_key = _safe_text(old_flow_key)
        new_key = _safe_text(new_flow_key)
        if not old_key or not new_key or old_key == new_key:
            return
        with self._lock:
            existing = self.latest_by_flow.get(old_key)
            if existing is None:
                return
            self.latest_by_flow[new_key] = existing
            self.latest_by_flow.pop(old_key, None)

    def _normalize_log_error(self, value):
        text = _safe_text(value)
        return text if text else ""

    def _append_execution_log(self, event, payload):
        if not event:
            return
        record = {
            "ts": _iso_now(),
            "event": str(event),
        }
        if isinstance(payload, dict):
            record.update(payload)
        try:
            line = json.dumps(record, ensure_ascii=False, default=str)
        except Exception:
            return
        with self._lock:
            with self._execution_log_path.open("a", encoding="utf-8") as handle:
                handle.write(line + "\n")

    def _register_flow_token(self, path):
        normalized = str(Path(path).resolve())
        for token, saved_path in self._flow_tokens.items():
            if saved_path == normalized:
                return token
        token = f"flow:{uuid.uuid4().hex}"
        self._flow_tokens[token] = normalized
        return token

    def _build_default_flow_file_name(self, mode):
        extension = MODE_EXTENSIONS.get(mode, ".zizd")
        return f"フロー{extension}"

    def _resolve_mode(self, config, file_name):
        metadata = config.get("metadata") if isinstance(config, dict) else {}
        raw_mode = _safe_text((metadata or {}).get("mode"))
        if raw_mode in MODE_EXTENSIONS:
            return raw_mode
        suffix = Path(file_name).suffix.lower()
        for mode, extension in MODE_EXTENSIONS.items():
            if extension == suffix:
                return mode
        return "workflow"

    def _is_secret_key(self, key):
        return _safe_text(key) in SECRET_FIELD_KEYS

    def _hide_sensitive_values(self, value, current_step="global"):
        if isinstance(value, list):
            return [self._hide_sensitive_values(item, current_step=current_step) for item in value]
        if isinstance(value, dict):
            next_step = _safe_text(value.get("step_id")) or current_step
            output = {}
            for key, item in value.items():
                if self._is_secret_key(key) and isinstance(item, (str, int, float)) and _safe_text(item):
                    ref, _ = self._store_hidden_value(
                        step_name=next_step,
                        field_key=_safe_text(key),
                        actual_value=str(item),
                    )
                    output[key] = ref
                    continue
                output[key] = self._hide_sensitive_values(item, current_step=next_step)
            return output
        return value

    def _restore_hidden_values(self, value):
        if isinstance(value, list):
            return [self._restore_hidden_values(item) for item in value]
        if isinstance(value, dict):
            return {key: self._restore_hidden_values(item) for key, item in value.items()}
        if _is_hidden_ref(value):
            return self.hidden_values.get(str(value), value)
        return value

    def _store_hidden_value(self, *, step_name, field_key, actual_value, current_ref=None):
        ref = current_ref if current_ref and current_ref in self.hidden_values else self._allocate_hidden_ref(step_name)
        self.hidden_values[ref] = actual_value
        self.hidden_meta[ref] = self._build_hidden_meta(field_key, actual_value)
        return ref, self.hidden_meta[ref]

    def _allocate_hidden_ref(self, step_name):
        scope = _sanitize_hidden_scope(step_name)
        next_index = self._hidden_counters.get(scope, 0) + 1
        self._hidden_counters[scope] = next_index
        return f"{{{{hidden.{scope}.var{next_index}}}}}"

    def _build_hidden_meta(self, field_key, actual_value):
        text = _safe_text(actual_value)
        key = _safe_text(field_key)
        if key in {"file_path", "output_path"}:
            path_obj = Path(text)
            return {
                "display_name": path_obj.name or text,
                "display_hint": path_obj.parent.name if path_obj.parent and path_obj.parent.name else "",
            }
        if key in {"folder_path", "directory", "output_folder", "output_dir"}:
            path_obj = Path(text)
            folder_name = path_obj.name or text
            return {
                "display_name": folder_name,
                "display_hint": path_obj.parent.name if path_obj.parent and path_obj.parent.name else "",
            }
        return {
            "display_name": key,
            "display_hint": "configured",
        }

    def _collect_secret_values(self, value):
        collected = set()
        if isinstance(value, list):
            for item in value:
                collected.update(self._collect_secret_values(item))
            return collected
        if isinstance(value, dict):
            for key, item in value.items():
                if self._is_secret_key(key) and isinstance(item, (str, int, float)) and _safe_text(item):
                    collected.add(str(item))
                collected.update(self._collect_secret_values(item))
        return collected

    def mask_log_message(self, run_id, message):
        with self._lock:
            session = self.runs.get(run_id) or {}
        masked = str(message or "")
        mask_targets = set()
        for item in session.get("secret_values") or set():
            text = _safe_text(item)
            if not text:
                continue
            mask_targets.add(text)
            mask_targets.add(text.replace("\\", "/"))
            mask_targets.add(text.replace("/", "\\"))
        for target in sorted(mask_targets, key=len, reverse=True):
            masked = masked.replace(target, "***********")
        return masked

    def _success_response(self, message_id, message_type, payload):
        return {
            "v": self.PROTOCOL_VERSION,
            "kind": "res",
            "type": message_type,
            "id": message_id,
            "ts": _iso_now(),
            "payload": payload,
        }

    def _error_response(self, message_id, message_type, code, message):
        return {
            "v": self.PROTOCOL_VERSION,
            "kind": "res",
            "type": message_type,
            "id": message_id,
            "ts": _iso_now(),
            "error": {
                "code": code,
                "message": message,
                "detail": {},
                "retryable": False,
                "trace_id": uuid.uuid4().hex,
            },
        }


class WebViewBridge(QObject):
    messageToFrontend = Signal(str)

    def __init__(self, runtime):
        super().__init__()
        self._runtime = runtime
        self._runtime.set_event_sink(self._emit_event)

    def _emit_event(self, message):
        self.messageToFrontend.emit(json.dumps(message, ensure_ascii=False))

    @Slot(str)
    def postMessage(self, raw_text):
        try:
            response = self._runtime.handle_message(raw_text)
        except ValueError as error:
            response = self._runtime._error_response(None, "unknown", "E_VALIDATION", str(error))
        except Exception:
            response = self._runtime._error_response(None, "unknown", "E_INTERNAL", "内部エラーが発生しました。")
        self.messageToFrontend.emit(json.dumps(response, ensure_ascii=False))
